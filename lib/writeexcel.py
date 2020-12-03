#!/usr/bin/env python
# _*_ coding:utf-8 _*_
__author__ = 'YinJia'

import os, sys

sys.path.append(os.path.dirname(os.path.dirname(__file__)))
import shutil
from config import setting
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import RED, GREEN, DARKYELLOW
import configparser as cparser

# --------- 读取config.ini配置文件 ---------------
cf = cparser.ConfigParser()
cf.read(setting.TEST_CONFIG, encoding='UTF-8')
name = cf.get("tester", "name")


class WriteExcel():
    """文件写入数据"""

    def __init__(self, fileName):
        self.filename = fileName
        if not os.path.exists(self.filename):
            # 文件不存在，则拷贝模板文件至指定报告目录下
            shutil.copyfile(setting.SOURCE_FILE, setting.TARGET_FILE)
        self.wb = load_workbook(self.filename)
        self.ws = self.wb.active

    def write_data(self, row_n, value, re_t, re_d):
        """
        写入测试结果
        :param row_n:数据所在行数
        :param value: 测试结果值
        :return: 无
        """
        font_GREEN = Font(name='宋体', color=GREEN, bold=True)
        font_RED = Font(name='宋体', color=RED, bold=True)
        font1 = Font(name='宋体', color=DARKYELLOW, bold=True)
        align = Alignment(horizontal='center', vertical='center')
        # 获数所在行数
        L_n = "L" + str(row_n)
        M_n = "M" + str(row_n)
        if value == "PASS":
            self.ws.cell(row_n, 12, value)
            self.ws[L_n].font = font_GREEN
        if value == "FAIL":
            self.ws.cell(row_n, 12, value)
            self.ws[L_n].font = font_RED
            self.ws.cell(row_n, 13, re_t)
            self.ws.cell(row_n, 14, re_d)
            self.ws[L_n].font = font_RED
        # self.ws.cell(row_n, 14, name)
        self.ws[L_n].alignment = align
        self.ws[M_n].font = font1
        self.ws[M_n].alignment = align
        self.wb.save(self.filename)

    def write_case(self):
        cases = (get_all_case(info) for info in apiInfo)
        for row, case in enumerate(cases):
            for idx in range(11):
                self.ws.cell(row + 2, idx + 1, case[idx])
        self.wb.save(self.filename)


if __name__ == '__main__':
    sys.path.append("C:/Users/shining3d/PycharmProjects/ApiDiff-py")
    from getapi import apiInfo, get_all_case
    from config.setting import SOURCE_FILE

    WriteExcel(SOURCE_FILE).write_case()
